from __future__ import annotations
import base64
import importlib.util
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from factory.engine import SupabaseClient


def _common() -> Any:
    path = Path(__file__).resolve().parents[1] / "_common.py"
    spec = importlib.util.spec_from_file_location("multi_shopper_common", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _storage_service() -> Any:
    path = Path(__file__).resolve().parents[2] / "vertical_supabase" / "supabase_storage_upload" / "service.py"
    spec = importlib.util.spec_from_file_location("supabase_storage_upload_service", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _price_compare(db: SupabaseClient, ctx: dict, products: list[dict[str, Any]], supplier_name: str | None) -> list[dict[str, Any]]:
    canonical_names = [str(item.get("raw_description") or "").strip() for item in products if str(item.get("raw_description") or "").strip()]
    if not canonical_names:
        return []

    result = db.rest_select(
        "price_history",
        filters={"company_id": ctx["company_id"], "product_name": canonical_names[0], "currency": ctx.get("currency") or "MXN"},
        select="unit_price,supplier_name,product_name,price_date",
        order="price_date.desc",
        limit=50,
    )
    if not result.get("ok"):
        return []
    rows = result.get("data") or []

    by_product: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        by_product.setdefault(str(row.get("product_name") or "").strip(), []).append(row)

    out: list[dict[str, Any]] = []
    for item in products:
        raw = str(item.get("raw_description") or "").strip()
        if not raw:
            continue
        prices = by_product.get(raw, [])
        out.append(
            {
                "canonical_name": raw,
                "unit_price_requested": item.get("unit_price"),
                "currency": item.get("currency") or ctx.get("currency") or "MXN",
                "supplier_name": supplier_name,
                "history_count": len(prices),
                "min_price": min((p.get("unit_price") for p in prices if p.get("unit_price") is not None), default=None),
                "max_price": max((p.get("unit_price") for p in prices if p.get("unit_price") is not None), default=None),
                "last_price": prices[0].get("unit_price") if prices else None,
                "last_supplier": prices[0].get("supplier_name") if prices else None,
                "last_price_date": prices[0].get("price_date") if prices else None,
            }
        )
    return out


class PurchaseImportSkillService:
    def ejecutar(self, context: dict) -> dict:
        ctx_result = _common().resolve_context(context)
        if not ctx_result.get("ok"):
            return ctx_result
        ctx = ctx_result["data"]
        dry_run = context.get("dry_run", True)
        action = context.get("action") or "import_document"

        if action == "compare":
            extracted = context.get("extracted") or {}
            products = extracted.get("products") or context.get("products") or []
            supplier_name = extracted.get("supplier_name") or context.get("supplier_name")
            return {"ok": True, "data": {"comparison": _price_compare(SupabaseClient(ctx), ctx, products, supplier_name)}}

        if action == "import_document":
            return self._import_document(context, ctx, dry_run)

        return {"ok": False, "error": f"action no soportada: {action}"}

    def _import_document(self, context: dict, ctx: dict, dry_run: bool) -> dict:
        required = {"file_name", "content_b64"}
        missing = [k for k in required if not context.get(k)]
        if missing:
            return {"ok": False, "error": f"campos requeridos: {', '.join(missing)}"}
        bucket = context.get("bucket") or context.get("storage_bucket")
        if not bucket:
            return {"ok": False, "error": "bucket/storage_bucket requerido"}

        upload = _storage_service().StorageUploadService().ejecutar(
            {
                "bucket": bucket,
                "path": context.get("storage_path") or f"{ctx['module_code']}/{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{self._safe_name(context.get('file_name'))}",
                "content_b64": context.get("content_b64"),
                "content_type": context.get("content_type") or context.get("file_type") or "application/octet-stream",
            }
        )
        if not upload.get("ok"):
            return upload
        storage = upload.get("data") or {}

        extracted_text = ""
        if context.get("content_b64"):
            try:
                extracted_text = self._extract_text(context.get("content_b64", ""), context.get("file_type", ""), context.get("file_name", ""))
            except Exception:
                extracted_text = ""

        extracted_data = context.get("extracted_data")
        if not extracted_data and extracted_text:
            extracted_data = self._run_ai_extract(ctx, extracted_text, context.get("file_type"), context.get("file_name"))

        doc_payload = {
            **ctx,
            "file_name": context.get("file_name"),
            "file_url": storage.get("url"),
            "storage_bucket": bucket,
            "storage_path": storage.get("path"),
            "file_type": context.get("file_type"),
            "extracted_text": extracted_text or None,
            "extracted_data": extracted_data or {},
            "source_context": {"source": "purchase_import_skill"},
            "processing_status": "extracted" if extracted_data else "pending",
            "action": "create",
        }
        if context.get("supplier_name"):
            doc_payload["related_supplier_name"] = context.get("supplier_name")

        doc_result = _common().MultiShopperCrudService("documents").ejecutar(doc_payload)
        document_id = doc_result.get("data", {}).get("id") if isinstance(doc_result.get("data"), dict) else None
        if isinstance(doc_result.get("data"), dict):
            document_id = doc_result["data"].get("id") or doc_result["data"].get("document", {}).get("id")

        products_out: list[dict[str, Any]] = []
        suppliers_out: list[dict[str, Any]] = []
        prices_out: list[dict[str, Any]] = []
        errors_out: list[dict[str, Any]] = []

        if extracted_data:
            products = (extracted_data.get("products") or []) or []
            supplier_name = str(extracted_data.get("supplier_name") or context.get("supplier_name") or "").strip() or None
            currency = str(extracted_data.get("currency") or context.get("currency") or "MXN").strip()
            document_date = str(extracted_data.get("document_date") or context.get("document_date") or "").strip() or None

            if supplier_name:
                sup = self._find_or_create(ctx, "suppliers", "name", supplier_name, {"status": "active"}, dry_run)
                if sup.get("ok"):
                    row = sup["data"]
                    suppliers_out.append({"name": supplier_name, "id": row.get("id"), "folio": row.get("folio"), "action": row.get("_action")})
                else:
                    errors_out.append({"item": supplier_name, "error": sup.get("error")})
            supplier_id = (suppliers_out[0].get("id") if suppliers_out else None)

            for item in products:
                raw = str(item.get("raw_description") or "").strip()
                if not raw:
                    continue
                prod = self._find_or_create(ctx, "products", "canonical_name", raw, {
                    "category_name": item.get("category_name"),
                    "unit": item.get("unit"),
                    "brand": item.get("brand"),
                    "status": "active",
                }, dry_run)
                if prod.get("ok"):
                    prow = prod["data"]
                    products_out.append({"canonical_name": raw, "id": prow.get("id"), "folio": prow.get("folio"), "action": prow.get("_action")})
                else:
                    errors_out.append({"item": raw, "error": prod.get("error")})
                    continue
                if supplier_id:
                    pp = self._upsert_price_history(ctx, supplier_name, supplier_id, item, raw, currency, document_date, document_id, dry_run)
                    if pp.get("ok"):
                        prices_out.extend(pp.get("data", {}).get("prices", []) if isinstance(pp.get("data"), dict) else [pp.get("data")])
                    else:
                        errors_out.append({"item": f"precio {raw}", "error": pp.get("error")})

        comparison = _price_compare(SupabaseClient(ctx), ctx, (extracted_data.get("products") if isinstance(extracted_data, dict) else []) or [], (extracted_data.get("supplier_name") if isinstance(extracted_data, dict) else None) or context.get("supplier_name"))

        result_payload = {
            "ok": True,
            "dry_run": dry_run,
            "document": doc_result.get("data"),
            "suppliers": suppliers_out,
            "products": products_out,
            "prices": prices_out,
            "comparison": comparison,
            "errors": errors_out,
            "summary": {
                "suppliers_count": len(suppliers_out),
                "products_created": len([p for p in products_out if p.get("action") in (None, "found") and (not dry_run)]),
                "prices_created": len(prices_out),
                "error_count": len(errors_out),
            },
        }
        return {"ok": True, "data": result_payload}

    def _run_ai_extract(self, ctx: dict, text: str, file_type: str | None, file_name: str | None) -> dict[str, Any] | None:
        try:
            path = Path(__file__).resolve().parents[2] / "vertical_factory_utils" / "ai_interpreter" / "service.py"
            spec = importlib.util.spec_from_file_location("ai_interpreter_service", path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            schema = {
                "supplier_name": None,
                "document_date": None,
                "currency": None,
                "products": [
                    {
                        "raw_description": None,
                        "quantity": None,
                        "unit": None,
                        "unit_price": None,
                        "subtotal": None,
                        "brand": None,
                        "category_name": None,
                    }
                ],
                "summary": None,
            }
            ai = module.run(
                {
                    "mode": "extract",
                    "schema": schema,
                    "text": text,
                    "content_b64": "",
                    "media_type": file_type or "text/plain",
                    "context": (
                        "Extrae todos los productos listados en este documento de cotizacion. "
                        "Cada producto debe tener raw_description. No inventes precios ni cantidades. Si no puedes leer un campo usa null."
                    ),
                }
            )
            if ai.get("ok"):
                return ai.get("data", {}).get("extracted") or {}
        except Exception:
            pass
        return None

    def _extract_text(self, content_b64: str, mime: str, filename: str) -> str:
        data = base64.b64decode(content_b64)
        ext = Path(filename).suffix.lower()
        if ext == ".csv":
            return data.decode("utf-8", errors="ignore")
        if ext in {".xlsx", ".xls"}:
            try:
                import openpyxl
            except ImportError:
                return ""
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.sheetnames:
                lines.append(f"=== {sheet} ===")
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        lines.append("\t".join(cells))
            return "\n".join(lines)
        if ext == ".pdf":
            try:
                import pdfplumber
            except ImportError:
                return ""
            texts: list[str] = []
            with pdfplumber.open(BytesIO(data)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        texts.append(t)
            return "\n".join(texts)
        if mime.startswith("image/") or ext in {".png", ".jpg", ".jpeg", ".webp"}:
            try:
                from PIL import Image
                import pytesseract
            except Exception:
                return ""
            text = pytesseract.image_to_string(Image.open(BytesIO(data)))
            return text or ""
        return ""

    def _find_or_create(self, ctx: dict, table: str, match_field: str, match_value: str, extra: dict[str, Any], dry_run: bool) -> dict[str, Any]:
        db = SupabaseClient(ctx)
        result = db.rest_select(table, filters={"company_id": ctx["company_id"], match_field: match_value}, select="*", limit=1)
        if result.get("ok") and result.get("data"):
            row = result["data"][0]
            row["_action"] = "found"
            return {"ok": True, "data": row}
        if dry_run:
            return {"ok": True, "data": {"id": None, "folio": None, match_field: match_value, "_action": "dry_run"}}
        row = {
            "company_id": ctx["company_id"],
            "project_code": ctx["project_code"],
            "module_code": ctx["module_code"],
            match_field: match_value,
            **{k: v for k, v in extra.items() if v is not None},
        }
        ins = db.rest_insert(table, [row])
        if not ins.get("ok"):
            return ins
        data = ins.get("data") or []
        saved = data[0] if isinstance(data, list) and data else data
        if isinstance(saved, dict):
            saved["_action"] = "created"
        return {"ok": True, "data": saved}

    def _upsert_price_history(self, ctx: dict, supplier_name: str | None, supplier_id: str | None, item: dict[str, Any], raw: str, currency: str, document_date: str | None, document_id: str | None, dry_run: bool) -> dict[str, Any]:
        unit_price = item.get("unit_price")
        product_id = (self._find_or_create(ctx, "products", "canonical_name", raw, {"status": "active"}, dry_run).get("data") or {}).get("id") if dry_run else None
        if not product_id:
            prod = self._find_or_create(ctx, "products", "canonical_name", raw, {"status": "active"}, dry_run)
            if not prod.get("ok"):
                return prod
            product_id = prod["data"].get("id")
        if unit_price is None or not supplier_id or not product_id:
            if dry_run:
                return {"ok": True, "data": {"preview_dry_run": {"product": raw, "unit_price": unit_price}}}
            return {"ok": False, "error": "unit_price/supplier/product requeridos para price_history"}
        row = {
            "company_id": ctx["company_id"],
            "project_code": ctx["project_code"],
            "module_code": ctx["module_code"],
            "product_id": product_id,
            "supplier_id": supplier_id,
            "product_name": raw,
            "supplier_name": supplier_name,
            "raw_description": raw,
            "unit_price": unit_price,
            "currency": currency,
            "price_type": "supplier_cost",
        }
        if item.get("quantity") is not None:
            row["quantity"] = item["quantity"]
        if item.get("unit"):
            row["unit"] = item["unit"]
        if item.get("subtotal") is not None:
            row["subtotal"] = item["subtotal"]
        if document_date:
            row["price_date"] = document_date
        if document_id:
            row["source_document_id"] = document_id
        if dry_run:
            return {"ok": True, "data": {"prices": [row]}}
        ins = SupabaseClient(ctx).rest_insert("price_history", [row])
        if not ins.get("ok"):
            return ins
        data = ins.get("data") or []
        saved = data[0] if isinstance(data, list) and data else data
        return {"ok": True, "data": {"prices": [saved] if isinstance(saved, dict) else []}}

    def _safe_name(self, filename: str) -> str:
        return re.sub(r"[^\w.\-]", "_", filename or "documento")
