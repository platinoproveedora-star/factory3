from __future__ import annotations

import base64
import importlib.util
import io
import re
from datetime import datetime
from pathlib import Path

from factory.engine import SupabaseClient


def _load(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _common():
    return _load(Path(__file__).resolve().parents[1] / "_common.py", "multi_shopper_common")


_XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/xlsx",
    "application/xls",
}

_EXTRACT_SCHEMA = {
    "supplier_name": None,
    "document_date": None,
    "currency": None,
    "products": [
        {
            "raw_description": None,
            "quantity": None,
            "unit": None,
            "unit_price": None,
            "subtotal": None,
            "brand": None,
            "category_name": None,
        }
    ],
    "summary": None,
}


def _xlsx_to_text(content_b64: str) -> str:
    try:
        import openpyxl
    except ImportError:
        return ""
    raw = base64.b64decode(content_b64)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                lines.append("\t".join(cells))
    return "\n".join(lines)


class DocumentSkillService:
    def ejecutar(self, context: dict) -> dict:
        action = context.get("action") or "list"
        if action == "extract":
            return self.extract(context)
        if action == "import_products":
            return self.import_products(context)
        if action == "create" and context.get("content_b64"):
            return self.upload_and_create(context)
        return _common().MultiShopperCrudService("documents").ejecutar(context)

    def upload_and_create(self, context: dict) -> dict:
        ctx_result = _common().resolve_context(context)
        if not ctx_result.get("ok"):
            return ctx_result
        ctx = ctx_result["data"]
        bucket = context.get("bucket") or context.get("storage_bucket")
        if not bucket:
            return {"ok": False, "error": "bucket/storage_bucket requerido para upload"}

        file_name = context.get("file_name") or context.get("filename") or "documento"
        safe_name = re.sub(r"[^\w.\-]", "_", file_name)
        path = context.get("storage_path") or f"{ctx['module_code']}/{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{safe_name}"

        storage_service = _load(
            Path(__file__).resolve().parents[2] / "vertical_supabase" / "supabase_storage_upload" / "service.py",
            "supabase_storage_upload_service",
        )
        uploaded = storage_service.StorageUploadService().ejecutar(
            {
                "bucket": bucket,
                "path": path,
                "content_b64": context.get("content_b64"),
                "content_type": context.get("file_type") or context.get("content_type") or "application/octet-stream",
            }
        )
        if not uploaded.get("ok"):
            return uploaded

        storage_data = uploaded.get("data") or {}

        # Para xlsx: extraer texto y guardarlo en extracted_text
        extracted_text = None
        file_type = context.get("file_type") or context.get("content_type") or ""
        is_xlsx = file_type in _XLSX_TYPES or file_name.lower().endswith((".xlsx", ".xls"))
        if is_xlsx:
            extracted_text = _xlsx_to_text(context.get("content_b64") or "")

        payload = {
            **ctx,
            "file_name": file_name,
            "file_url": storage_data.get("url"),
            "storage_bucket": bucket,
            "storage_path": storage_data.get("path"),
            "action": "create",
        }
        if extracted_text:
            payload["extracted_text"] = extracted_text

        result = _common().MultiShopperCrudService("documents").ejecutar(payload)
        if result.get("ok"):
            result["data"]["can_extract"] = bool(
                is_xlsx or file_type == "application/pdf" or (file_type or "").startswith("image/")
            )
        return result

    def extract(self, context: dict) -> dict:
        """Extrae productos de un documento guardado (por id/folio) o de content_b64 directo."""
        ctx_result = _common().resolve_context(context)
        if not ctx_result.get("ok"):
            return ctx_result
        ctx = ctx_result["data"]

        text = context.get("text") or context.get("extracted_text") or ""
        content_b64 = context.get("content_b64") or ""
        media_type = context.get("media_type") or context.get("file_type") or "text/plain"
        doc_id = None

        # Si se pasa id o folio, buscar el documento en la DB
        if not text and not content_b64:
            doc_id = context.get("id") or context.get("document_id")
            folio = context.get("folio")
            if not doc_id and not folio:
                return {"ok": False, "error": "Se requiere id, folio, content_b64 o text"}
            filters = {"company_id": ctx["company_id"]}
            if doc_id:
                filters["id"] = doc_id
            else:
                filters["folio"] = folio
            db_result = SupabaseClient(ctx).rest_select("documents", filters=filters, select="*", limit=1)
            if not db_result.get("ok"):
                return db_result
            rows = db_result.get("data") or []
            if not rows:
                return {"ok": False, "error": "Documento no encontrado"}
            doc = rows[0]
            doc_id = doc.get("id")
            text = doc.get("extracted_text") or ""
            media_type = doc.get("file_type") or "text/plain"
            if not text:
                return {"ok": False, "error": "El documento no tiene texto extraido. Solo se soporta extraccion de xlsx, PDF e imagenes."}

        ai_service = _load(
            Path(__file__).resolve().parents[2] / "vertical_factory_utils" / "ai_interpreter" / "service.py",
            "ai_interpreter_service",
        )
        ai_result = ai_service.run(
            {
                "mode": "extract",
                "schema": _EXTRACT_SCHEMA,
                "text": text,
                "content_b64": content_b64,
                "media_type": media_type,
                "context": (
                    "Extrae todos los productos listados en este documento de cotizacion. "
                    "Cada producto debe tener raw_description. "
                    "No inventes precios ni cantidades. Si no puedes leer un campo usa null."
                ),
            }
        )
        if not ai_result.get("ok"):
            return ai_result

        extracted = ai_result.get("data", {}).get("extracted") or {}

        # Guardar extracted_data en la tabla documents y marcar como procesado
        if doc_id and not context.get("dry_run", False):
            import json
            SupabaseClient(ctx).rest_update(
                "documents",
                {
                    "extracted_data": json.dumps(extracted),
                    "processing_status": "extracted",
                },
                {"id": doc_id, "company_id": ctx["company_id"]},
            )

        return {"ok": True, "data": {"extracted": extracted, "document_id": doc_id}}

    def import_products(self, context: dict) -> dict:
        """
        A partir de la extraccion de un documento, guarda en DB:
          1. Proveedor (find-or-create por nombre)
          2. Productos (find-or-create por canonical_name)
          3. price_history por cada producto con precio, enlazado al documento fuente
        """
        ctx_result = _common().resolve_context(context)
        if not ctx_result.get("ok"):
            return ctx_result
        ctx = ctx_result["data"]

        products = context.get("products") or []
        if not products:
            return {"ok": False, "error": "products requerido"}

        dry_run = context.get("dry_run", True)
        document_id = context.get("document_id") or context.get("id") or None
        supplier_name = str(context.get("supplier_name") or "").strip() or None
        currency = str(context.get("currency") or "MXN").strip() or "MXN"
        price_date = str(context.get("document_date") or "").strip() or None

        db = SupabaseClient(ctx)

        # 1 — Find or create supplier
        supplier_id = None
        supplier_folio = None
        supplier_action = None
        if supplier_name:
            sup_result = _find_or_create(
                db, ctx,
                table="suppliers",
                match_field="name",
                match_value=supplier_name,
                extra={"status": "active"},
                dry_run=dry_run,
            )
            if sup_result.get("ok"):
                supplier_id = sup_result["data"].get("id")
                supplier_folio = sup_result["data"].get("folio")
                supplier_action = sup_result["data"].get("_action")

        # 2 — For each product: find-or-create + price_history
        products_created = []
        products_found = []
        prices_created = []
        errors = []

        for item in products:
            raw = str(item.get("raw_description") or "").strip()
            if not raw:
                continue

            canonical = raw
            # find-or-create product
            prod_result = _find_or_create(
                db, ctx,
                table="products",
                match_field="canonical_name",
                match_value=canonical,
                extra={
                    "category_name": item.get("category_name") or None,
                    "unit": item.get("unit") or None,
                    "brand": item.get("brand") or None,
                    "status": "active",
                },
                dry_run=dry_run,
            )
            if not prod_result.get("ok"):
                errors.append({"item": raw, "error": prod_result.get("error")})
                continue

            prod = prod_result["data"]
            product_id = prod.get("id")
            if prod.get("_action") == "created":
                products_created.append({"folio": prod.get("folio"), "canonical_name": canonical})
            else:
                products_found.append({"folio": prod.get("folio"), "canonical_name": canonical})

            # 3 — price_history si hay precio
            unit_price = item.get("unit_price")
            if unit_price is not None and supplier_id and product_id and not dry_run:
                ph_row = {
                    "company_id": ctx["company_id"],
                    "project_code": ctx["project_code"],
                    "module_code": ctx["module_code"],
                    "product_id": product_id,
                    "supplier_id": supplier_id,
                    "product_name": canonical,
                    "supplier_name": supplier_name,
                    "raw_description": raw,
                    "unit_price": unit_price,
                    "currency": currency,
                    "price_type": "supplier_cost",
                }
                if item.get("quantity") is not None:
                    ph_row["quantity"] = item["quantity"]
                if item.get("unit"):
                    ph_row["unit"] = item["unit"]
                if item.get("subtotal") is not None:
                    ph_row["subtotal"] = item["subtotal"]
                if price_date:
                    ph_row["price_date"] = price_date
                if document_id:
                    ph_row["source_document_id"] = document_id
                ph_ins = db.rest_insert("price_history", [ph_row])
                if ph_ins.get("ok"):
                    ph_data = ph_ins.get("data") or []
                    ph_saved = ph_data[0] if isinstance(ph_data, list) and ph_data else ph_data
                    prices_created.append({"folio": ph_saved.get("folio") if isinstance(ph_saved, dict) else None, "product": canonical, "price": unit_price})
                else:
                    errors.append({"item": f"precio {raw}", "error": ph_ins.get("error")})

        return {
            "ok": True,
            "data": {
                "dry_run": dry_run,
                "supplier": {"id": supplier_id, "folio": supplier_folio, "name": supplier_name, "action": supplier_action},
                "products_created": len(products_created),
                "products_found": len(products_found),
                "prices_created": len(prices_created),
                "errors": len(errors),
                "detail": {
                    "products_created": products_created,
                    "products_found": products_found,
                    "prices": prices_created,
                    "error_detail": errors,
                },
            },
        }


def _find_or_create(db: SupabaseClient, ctx: dict, table: str, match_field: str, match_value: str, extra: dict, dry_run: bool) -> dict:
    """Busca un registro por match_field=match_value; si no existe lo crea."""
    result = db.rest_select(table, filters={"company_id": ctx["company_id"], match_field: match_value}, select="*", limit=1)
    if result.get("ok") and result.get("data"):
        row = result["data"][0]
        row["_action"] = "found"
        return {"ok": True, "data": row}

    if dry_run:
        return {"ok": True, "data": {"id": None, "folio": None, match_field: match_value, "_action": "dry_run"}}

    row = {
        "company_id": ctx["company_id"],
        "project_code": ctx["project_code"],
        "module_code": ctx["module_code"],
        match_field: match_value,
        **{k: v for k, v in extra.items() if v is not None},
    }
    ins = db.rest_insert(table, [row])
    if not ins.get("ok"):
        return ins
    data = ins.get("data") or []
    saved = data[0] if isinstance(data, list) and data else data
    if isinstance(saved, dict):
        saved["_action"] = "created"
    return {"ok": True, "data": saved}
