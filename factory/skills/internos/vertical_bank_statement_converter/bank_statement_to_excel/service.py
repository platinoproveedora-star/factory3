from __future__ import annotations

import base64
import io
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from _statement_common import resolve_statement_context
from factory.engine import SupabaseClient


class BankStatementToExcelService:
    def ejecutar(self, context: dict) -> dict:
        ctx_res = resolve_statement_context(context)
        if not ctx_res.get("ok"):
            return ctx_res
        ctx = ctx_res["data"]

        extraction_id = str(context.get("extraction_id") or "").strip()
        extraction_folio = str(context.get("extraction_folio") or "").strip()
        if not extraction_id and not extraction_folio:
            return {"ok": False, "error": "extraction_id o extraction_folio requerido"}

        db = SupabaseClient(ctx)

        filters: dict = {"empresa_id": f"eq.{ctx['company_id']}"}
        if extraction_id:
            filters["id"] = f"eq.{extraction_id}"
        else:
            filters["folio"] = f"eq.{extraction_folio}"

        ext_res = db.rest_select("statement_extractions", filters=filters, select="*", limit=1)
        if not ext_res.get("ok"):
            return ext_res
        if not ext_res.get("data"):
            return {"ok": False, "error": "extraccion no encontrada o no pertenece a esta empresa"}

        extraction = ext_res["data"][0]
        eid = extraction["id"]

        lines_res = db.rest_select(
            "statement_extracted_lines",
            filters={"extraction_id": f"eq.{eid}", "empresa_id": f"eq.{ctx['company_id']}"},
            select="*",
            order="raw_line_order.asc",
            limit=5000,
        )
        if not lines_res.get("ok"):
            return lines_res
        lines = lines_res.get("data") or []

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return {"ok": False, "error": "openpyxl no instalado. Agregar a requirements.txt"}

        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor="1F4E79")
        info_fill = PatternFill("solid", fgColor="D6E4F0")

        ws_mov = wb.active
        ws_mov.title = "Movimientos"

        period_str = f"{extraction.get('statement_period_start') or ''} - {extraction.get('statement_period_end') or ''}".strip(" -")
        info_rows = [
            ["Cuenta:", extraction.get("account_number_mask") or ""],
            ["CLABE:", extraction.get("clabe") or ""],
            ["Titular:", extraction.get("holder_name") or ""],
            ["Período:", period_str],
            ["Banco:", extraction.get("bank_name") or extraction.get("bank_profile") or ""],
            [],
        ]
        for row in info_rows:
            ws_mov.append(row)
            if row:
                ws_mov.cell(row=ws_mov.max_row, column=1).font = Font(bold=True)
                ws_mov.cell(row=ws_mov.max_row, column=1).fill = info_fill

        mov_headers = [
            "folio", "raw_line_order", "line_date", "direction", "amount", "saldo",
            "tipo_movimiento",
            "description",
            "clave_rastreo", "referencia",
            "nombre_origen", "cuenta_origen",
            "nombre_destino", "cuenta_destino",
            "banco_origen", "banco_destino", "rfc_origen", "rfc_destino",
            "hora_liquidacion", "concepto", "sucursal",
            "confidence", "parse_warnings",
        ]
        ws_mov.append(mov_headers)
        for cell in ws_mov[ws_mov.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for line in lines:
            meta = line.get("metadata") or {}
            banco_origen = " ".join(filter(None, [meta.get("banco_origen_codigo"), meta.get("banco_origen_nombre")])) or ""
            banco_destino = " ".join(filter(None, [meta.get("banco_destino_codigo"), meta.get("banco_destino_nombre")])) or ""
            ws_mov.append([
                line.get("folio"), line.get("raw_line_order"),
                str(line.get("line_date") or ""),
                line.get("direction"), line.get("amount"), line.get("saldo"),
                meta.get("tipo_movimiento"),
                line.get("description"),
                line.get("clave_rastreo"), line.get("referencia"),
                line.get("nombre_origen"), line.get("cuenta_origen"),
                line.get("nombre_destino"), line.get("cuenta_destino"),
                banco_origen, banco_destino,
                meta.get("rfc_origen"), meta.get("rfc_destino"),
                meta.get("hora_liquidacion"),
                meta.get("concepto"),
                meta.get("sucursal"),
                line.get("confidence"),
                json.dumps(line.get("parse_warnings") or [], ensure_ascii=False),
            ])

        ws_res = wb.create_sheet("Resumen")
        for row in info_rows:
            ws_res.append(row)
            if row:
                ws_res.cell(row=ws_res.max_row, column=1).font = Font(bold=True)
                ws_res.cell(row=ws_res.max_row, column=1).fill = info_fill

        res_headers = [
            "extraction_folio", "bank_profile", "profile_version", "bank_name",
            "holder_name", "clabe", "account_number_mask",
            "statement_period_start", "statement_period_end",
            "file_name", "file_hash", "storage_bucket", "storage_path",
            "validation_status", "status",
            "total_deposits_reported", "total_deposits_extracted", "validation_diff_deposits",
            "total_withdrawals_reported", "total_withdrawals_extracted", "validation_diff_withdrawals",
            "warnings", "total_lines_extracted",
        ]
        ws_res.append(res_headers)
        for cell in ws_res[ws_res.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        ws_res.append([
            extraction.get("folio"), extraction.get("bank_profile"), extraction.get("profile_version"),
            extraction.get("bank_name"), extraction.get("holder_name"), extraction.get("clabe"),
            extraction.get("account_number_mask"),
            str(extraction.get("statement_period_start") or ""), str(extraction.get("statement_period_end") or ""),
            extraction.get("file_name"), extraction.get("file_hash"),
            extraction.get("storage_bucket"), extraction.get("storage_path"),
            extraction.get("validation_status"), extraction.get("status"),
            extraction.get("total_deposits_reported"), extraction.get("total_deposits_extracted"),
            extraction.get("validation_diff_deposits"),
            extraction.get("total_withdrawals_reported"), extraction.get("total_withdrawals_extracted"),
            extraction.get("validation_diff_withdrawals"),
            json.dumps(extraction.get("warnings") or [], ensure_ascii=False),
            extraction.get("total_lines_extracted"),
        ])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_b64 = base64.b64encode(buf.getvalue()).decode()
        filename = f"{extraction.get('folio', 'estado')}_{extraction.get('bank_profile', 'bank')}.xlsx"

        return {
            "ok": True,
            "data": {
                "filename": filename,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "xlsx_base64": xlsx_b64,
                "lines_count": len(lines),
                "extraction_folio": extraction.get("folio"),
            },
        }
